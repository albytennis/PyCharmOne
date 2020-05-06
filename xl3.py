from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "ALBY"
cell_range = ws['A1':'C2']
print(cell_range)
colC      = ws['C']
print("colC:",colC)
col_range = ws['C:E']
print("col_range:",col_range)
row10     = ws[10]
row_range = ws[5:10]
for row in ws.iter_rows(min_row=1,max_row=2,max_col=3):
    for cell in row:
        print("ws.iter_rows transcends all cols in a row:",cell)

for col in ws.iter_cols(min_row=1,max_col=3,max_row=2):
    for cell in col:
        print("ws.iter_col transcends a column by row: ",cell)
ws['A1']= "First sheet"
ws2 = wb.create_sheet("Sheet2")

ws2['C9'] = 'hello world'
ws2.title = "Second Sheet"
print('tuple(ws2.rows:',tuple(ws2.rows))
print('tuple(ws.rows:',tuple(ws.rows))
print("'ws2['C9']=",ws2['C9'].value)
print("ws['A1']",ws['A1'].value)
ws2 = wb.active
wb.save("sample3.xlsx")