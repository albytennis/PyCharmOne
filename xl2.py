from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet("ws1")
ws2 = wb.create_sheet("ws2")
ws3 = wb.create_sheet("ws3")
ws.title = "YA TITLE"
ws.sheet_properties.tabColor = "1072BA"
print(wb.sheetnames)
for sheet in wb:
    print(sheet.title)
wb.save("exam2.xlsx")