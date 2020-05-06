# xl1.py   can we get to Excel
# pip install openpyxl
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1']  = 42
# rows can be appended
ws.append([1, 2, 3,"Hong","Vinny"])
ws.append(["Alex","Sam","Robkin"])
#python types will automatically be converted
import datetime
ws['A10']  = datetime.datetime.now()

#save file
wb.save("vinny.xlsx")


